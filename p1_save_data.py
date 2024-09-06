from utils import *
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook

cwd = os.getcwd()

# Define file paths
template_file = os.path.join(cwd, 'template', 'result1.xlsx')
output_file = os.path.join(cwd, 'result1.xlsx')

# Delete the output file if it already exists and create a new blank Excel file
if os.path.exists(output_file):
    os.remove(output_file)
pd.DataFrame().to_excel(output_file)

# Load the template Excel file
wb = load_workbook(template_file)

# Initialize data arrays
nodes = 224
times = 301

all_positions = np.ndarray(shape=(nodes * 2, times))
all_velocities = np.ndarray(shape=(nodes, times))
all_positions.fill(0)
all_velocities.fill(0)

# Calculate positions and velocities for each time step
for t in range(times):
    theta_s = find_theta_at_time(t)
    positions = calculate_spiral_positions(theta_s)
    velocities = calculate_velocities(positions)

    # Fill the arrays with position and velocity data
    for i in range(nodes):
        all_positions[2 * i][t] = np.round(positions[i][0], 6)  # x-coordinate
        all_positions[2 * i + 1][t] = np.round(positions[i][1], 6)  # y-coordinate
        all_velocities[i][t] = np.round(np.sqrt(velocities[i][0]**2 + velocities[i][1]**2), 6)  # velocity magnitude

# Write the position data to the "位置" sheet
ws_positions = wb["位置"]
for i in range(nodes * 2):
    for t in range(times):
        ws_positions.cell(row=i+2, column=t+2, value=all_positions[i][t])

# Write the velocity data to the "速度" sheet
ws_velocities = wb["速度"]
for i in range(nodes):
    for t in range(times):
        ws_velocities.cell(row=i+2, column=t+2, value=all_velocities[i][t])

# Save the modified workbook
wb.save(output_file)

print(f"Data saved to {output_file} successfully.")