from utils import *
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook

cwd = os.getcwd()

# Define file paths
template_file = os.path.join(cwd, 'template', 'result2.xlsx')
output_file = os.path.join(cwd, 'result2.xlsx')

# Delete the output file if it already exists and create a new blank Excel file
if os.path.exists(output_file):
    os.remove(output_file)
pd.DataFrame().to_excel(output_file)

# Load the template Excel file
wb = load_workbook(template_file)

# t = 362.993657
t = 412.473837

theta_s = find_theta_at_time(t)
positions = calculate_spiral_positions(theta_s)
velocities = calculate_velocities(positions)

result2 = np.ndarray(shape=(224, 3))
result2.fill(0)

# Fill the arrays with position and velocity data
for i in range(224):
    result2[i][0] = np.round(positions[i][0], 6)  # x-coordinate
    result2[i][1] = np.round(positions[i][1], 6)  # y-coordinate
    result2[i][2] = np.round(np.sqrt(velocities[i][0]**2 + velocities[i][1]**2), 6)  # velocity magnitude

# Write the position data to the "Sheet1" sheet
ws_result = wb["Sheet1"]
for i in range(224):
    for j in range(3):
        ws_result.cell(row=i+2, column=j+2, value=result2[i][j])

# Save the modified workbook
wb.save(output_file)
