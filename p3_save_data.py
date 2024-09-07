from utils import *
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook

cwd = os.getcwd()

# Define file paths
template_file = os.path.join(cwd, 'template', 'result2.xlsx')
output_file = os.path.join(cwd, 'p3.xlsx')

# Delete the output file if it already exists and create a new blank Excel file
if os.path.exists(output_file):
    os.remove(output_file)
pd.DataFrame().to_excel(output_file)

# Load the template Excel file
wb = load_workbook(template_file)

# theta_s = 56.81433157908003
# pitch = 0.497662
theta_s = 67.78172662837748
pitch = 0.417138
positions = calculate_spiral_positions(theta_s, pitch=pitch)
velocities = calculate_velocities(positions)

result3 = np.ndarray(shape=(224, 3))
result3.fill(0)

# Fill the arrays with position and velocity data
for i in range(224):
    result3[i][0] = np.round(positions[i][0], 6)  # x-coordinate
    result3[i][1] = np.round(positions[i][1], 6)  # y-coordinate
    result3[i][2] = np.round(np.sqrt(velocities[i][0]**2 + velocities[i][1]**2), 6)  # velocity magnitude

# Write the position data to the "Sheet1" sheet
ws_result = wb["Sheet1"]
for i in range(224):
    for j in range(3):
        ws_result.cell(row=i+2, column=j+2, value=result3[i][j])

# Save the modified workbook
wb.save(output_file)
